#! python3
# blankRowInserter.py - Copies blank lines into selected place in spreadsheet

import openpyxl, sys

usage = '''Usage: "python blankRowInserter.py N M example.xlsx"
        Where N is row after which to insert blank rows and M is the number of blank rows to insert
        E.g. python blankRowInserter.py 3 2 example.xlsx"'''

try:
    # N = row after which to insert blank rows
    N = int(sys.argv[1])
    # M = Number of blank rows to insert
    M = int(sys.argv[2])
    excelFile = sys.argv[3]
except:
    print(usage)
    sys.exit()
if len(sys.argv) != 4:
    print(usage)
    sys.exit()
if not excelFile.endswith('.xlsx'):
    print(usage)
    sys.exit()

# Open workbook to which blank rows will be added
wb = openpyxl.load_workbook(excelFile)
sheet = wb.active

# Open new workbook to save sheet with blank rows added
wbNew = openpyxl.Workbook()
sheetNew = wbNew.active

for row in sheet.rows:
    if N == 0:
        # Add M to the row number to print cells after blank rows
        for cell in row:
            newCell = sheetNew.cell(row=cell.row + M, column=cell.column)
            sheetNew[newCell.coordinate] = cell.value
        continue
    # Print out cells before blank rows
    for cell in row:
        sheetNew[cell.coordinate] = cell.value
    # Counter to count down (to 0) number of rows before blank rows are added
    N -= 1

# Save new spreadsheet to new file.
try:
    wbNew.save('blankRow_' + excelFile)
except PermissionError:
    print('Cannot save excel sheet while it is open. Close it.')
    sys.exit()

print('Success.')
